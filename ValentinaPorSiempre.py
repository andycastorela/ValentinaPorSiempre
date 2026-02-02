import streamlit as st
import pandas as pd
from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path
import base64
from supabase import create_client
from dotenv import load_dotenv
import os


# ==========================================================
#                 LOAD ENVIRONMENT VARIABLES
# ==========================================================
load_dotenv()
SUPABASE_URL = os.getenv("SUPABASE_URL", "https://uumezwowrtumbonsotyc.supabase.co")
SUPABASE_KEY = os.getenv(
    "SUPABASE_KEY",
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InV1bWV6d293cnR1bWJvbnNvdHljIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NjExNTk1MzUsImV4cCI6MjA3NjczNTUzNX0.dZGdfqa7BuYH6_W3yqirn8DsuEoffnyBm1qoLU-K0A0",
)
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)


# ==========================================================
#           ENSURE last_edit TABLE EXISTS
# ==========================================================
def ensure_last_edit_table():
    try:
        supabase.table("last_edit").select("*").limit(1).execute()
    except Exception:
        try:
            supabase.rpc("execute_sql", {
                "sql": """
                create table if not exists public.last_edit (
                    id bigint primary key,
                    user_name text,
                    timestamp timestamptz default now()
                );
                """
            }).execute()
        except Exception as e:
            st.warning(f"No se pudo verificar/crear la tabla last_edit: {e}")


ensure_last_edit_table()


# ==========================================================
#               PAGE CONFIGURATION
# ==========================================================
st.set_page_config(page_title="Valentina por Siempre", page_icon="VxS_logo.png", layout="wide")


# ==========================================================
#               LAST EDIT HELPERS
# ==========================================================
def update_last_edit(user_name):
    now = datetime.now().isoformat()
    try:
        supabase.table("last_edit").upsert({"id": 1, "user_name": user_name, "timestamp": now}).execute()
    except Exception as e:
        st.warning(f"⚠️ No se pudo actualizar el registro de edición en Supabase: {e}")


def get_last_edit():
    try:
        result = supabase.table("last_edit").select("*").eq("id", 1).execute()
        if result.data:
            record = result.data[0]
            return record.get("user_name"), record.get("timestamp")
    except Exception:
        return None, None
    return None, None


# ==========================================================
#                 ACCESS CONTROL (LOGIN)
# ==========================================================
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False
    st.session_state.user_name = ""


if not st.session_state.authenticated:
    st.sidebar.title("🔐 Ingreso")
    user_key = st.sidebar.text_input("Introduce tu clave de acceso:", type="password")


    AUTHORIZED_KEYS = {
        "equipo_vxs": None,
        "valentina_master": "Andrea"
    }


    if user_key not in AUTHORIZED_KEYS:
        st.warning("Por favor, introduce la clave de acceso para continuar.")
        st.stop()


    if AUTHORIZED_KEYS[user_key] is None:
        user_name = st.sidebar.text_input("Tu nombre (para registrar ediciones):")
        if not user_name:
            st.info("Por favor, escribe tu nombre para continuar.")
            st.stop()
    else:
        user_name = AUTHORIZED_KEYS[user_key]


    if st.sidebar.button("Ingresar"):
        st.session_state.authenticated = True
        st.session_state.user_name = user_name
        st.rerun()


# ==========================================================
#                 STYLES & LOGO
# ==========================================================
def load_logo_base64(path: str):
    file_path = Path(path)
    if file_path.exists():
        with open(file_path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    return None


logo_b64 = load_logo_base64("VxS_logo.png")
st.markdown(f"""
    <style>
    .main {{ background-color: #f7f5f2 !important; }}
    .custom-title {{
        text-align: center;
        color: #352208;
        font-size: 48px;
        font-weight: 800;
        margin-top: -5px;
        margin-bottom: 15px;
    }}
    .corner-image {{
        position: fixed; top: 80px; right: 25px;
        width: 90px; border-radius: 50%;
        box-shadow: 0px 4px 10px rgba(0,0,0,0.2);
        z-index: 999;
    }}
    .bottom-left {{
        position: fixed; bottom: 10px; left: 15px;
        color: #666; font-size: 14px;
        background-color: rgba(255,255,255,0.8);
        padding: 5px 10px; border-radius: 8px;
    }}
    /* Text wrap fix for tables */
    table.dataframe td, table.dataframe th {{
        white-space: normal !important;
        word-wrap: break-word !important;
        max-width: 300px !important;
        text-align: left !important;
        vertical-align: top !important;
        padding: 6px !important;
    }}
    </style>
    {'<img src="data:image/png;base64,' + logo_b64 + '" class="corner-image">' if logo_b64 else ''}
""", unsafe_allow_html=True)


st.markdown("<h1 class='custom-title'>💛 Valentina por Siempre</h1>", unsafe_allow_html=True)


# ==========================================================
#                 HELPER FUNCTIONS
# ==========================================================
def calculate_age(dob):
    if isinstance(dob, str):
        dob = datetime.strptime(dob, "%Y-%m-%d").date()
    today = date.today()
    return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))


def style_excel(df, filename):
    for col in ["fecha_nacimiento", "fecha_ultimo_apoyo"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce').dt.date
    df.to_excel(filename, index=False)
    wb = load_workbook(filename)
    ws = wb.active
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="ff8330")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    ws.freeze_panes = "A2"


    paliativos_fill = PatternFill("solid", fgColor="FFAB66")
    paliativos_col = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "cuidados_paliativos":
            paliativos_col = idx
            break
    if paliativos_col:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[paliativos_col - 1].value in (1, True, "1", "true", "True"):
                for cell in row:
                    cell.fill = paliativos_fill
    wb.save(filename)


def display_wrapped_table(df):
    """Display a wrapped HTML DataFrame with cuidados paliativos highlight."""
    df = df.copy()
    def row_style(row):
        style = ""
        if row.get("cuidados_paliativos") in [1, True, "1", "true", "True"]:
            style = 'background-color: #FFAB66;'
        return f'<tr style="{style}">' + ''.join(f"<td>{x}</td>" for x in row) + "</tr>"


    table_html = (
        "<table class='dataframe'>"
        "<thead><tr>" + "".join(f"<th>{c}</th>" for c in df.columns) + "</tr></thead>"
        "<tbody>" + "".join(row_style(row) for _, row in df.iterrows()) + "</tbody></table>"
    )
    st.markdown(table_html, unsafe_allow_html=True)


# ==========================================================
#                 MAIN INTERFACE
# ==========================================================
if st.session_state.authenticated:
    page = st.sidebar.radio(
        "Navegación",
        ["➕ Agregar Paciente", "📋 Ver / Editar Pacientes", "🎂 Cumpleaños"]
    )


    # ---------------- ADD PATIENT ----------------
    if page == "➕ Agregar Paciente":
        st.subheader("Ingresar nuevo paciente")
        with st.form("add_patient_form"):
            nombre = st.text_input("Nombre del paciente")
            fecha_nacimiento = st.date_input("Fecha de nacimiento", min_value=date(1900, 1, 1))
            nombre_tutor = st.text_input("Nombre del tutor")
            diagnostico = st.text_input("Diagnóstico")
            etapa_tratamiento = st.selectbox("Etapa del tratamiento", [
                "Diagnóstico inicial", "En tratamiento", "En vigilancia", "Cuidados paliativos"
            ])
            hospital = st.text_input("Hospital")
            estado_origen = st.text_input("Estado de origen")
            telefono_contacto = st.text_input("Celular de contacto")
            apoyos_entregados = st.text_input("Apoyos entregados")
            fecha_ultimo_apoyo = st.date_input("Fecha del último apoyo", value=None)
            notas = st.text_area("Notas")
            estado = st.selectbox("Estado del paciente", ["activo", "vigilancia", "fallecido"])
            cuidados_paliativos = st.checkbox("¿Está en cuidados paliativos?")
            submitted = st.form_submit_button("Agregar paciente")


            if submitted:
                data = {
                    "nombre": nombre,
                    "fecha_nacimiento": str(fecha_nacimiento),
                    "nombre_tutor": nombre_tutor,
                    "diagnostico": diagnostico,
                    "etapa_tratamiento": etapa_tratamiento,
                    "hospital": hospital,
                    "estado_origen": estado_origen,
                    "telefono_contacto": telefono_contacto,
                    "apoyos_entregados": apoyos_entregados,
                    "fecha_ultimo_apoyo": str(fecha_ultimo_apoyo) if fecha_ultimo_apoyo else None,
                    "notas": notas,
                    "estado": estado,
                    "cuidados_paliativos": cuidados_paliativos
                }
                supabase.table("pacientes").insert(data).execute()
                update_last_edit(st.session_state.user_name)
                st.success(f"✅ Paciente agregado exitosamente por {st.session_state.user_name}.")


    # ---------------- VIEW / EDIT / DELETE ----------------
    elif page == "📋 Ver / Editar Pacientes":
        st.subheader("❤️‍🩹 Lista de pacientes")


        query = supabase.table("pacientes").select("*").execute()
        df = pd.DataFrame(query.data)
        if not df.empty and "id" in df.columns:
            df = df.sort_values(by="id", ascending=True).reset_index(drop=True)
            df["Edad"] = df["fecha_nacimiento"].apply(calculate_age)


            display_wrapped_table(df)  #highlight cuidados paliativos


            # --- EDIT SECTION ---
            selected_id = st.selectbox("Selecciona ID del paciente para editar", df["id"].tolist())
            patient_data = df[df["id"] == selected_id].iloc[0]


            with st.form("edit_patient_form"):
                nombre_tutor = st.text_input("Nombre del tutor", value=patient_data["nombre_tutor"])
                diagnostico = st.text_input("Diagnóstico", value=patient_data["diagnostico"])
                hospital = st.text_input("Hospital", value=patient_data["hospital"])
                notas = st.text_area("Notas", value=patient_data["notas"])
                estado = st.selectbox("Estado del paciente", ["activo", "vigilancia", "fallecido"],
                                      index=["activo", "vigilancia", "fallecido"].index(patient_data["estado"]))
                submitted_edit = st.form_submit_button("💾 Guardar cambios")


                if submitted_edit:
                    update_data = {
                        "nombre_tutor": nombre_tutor,
                        "diagnostico": diagnostico,
                        "hospital": hospital,
                        "notas": notas,
                        "estado": estado
                    }
                    supabase.table("pacientes").update(update_data).eq("id", selected_id).execute()
                    update_last_edit(st.session_state.user_name)
                    st.success("✅ Cambios guardados correctamente.")
                    st.rerun()


            # --- DELETE SECTION ---
            delete_id = st.number_input("🗑️ ID del paciente a eliminar", min_value=0, step=1)
            if st.button("Confirmar eliminación"):
                st.warning(f"⚠️ ¿Seguro que deseas eliminar el paciente con ID {delete_id}?")
                if st.button("✅ Sí, eliminar permanentemente"):
                    supabase.table("pacientes").delete().eq("id", delete_id).execute()
                    update_last_edit(st.session_state.user_name)
                    st.success(f"🗑️ Paciente con ID {delete_id} eliminado correctamente.")
                    st.rerun()


            # --- EXPORT ---
            if st.button("📥 Exportar a Excel"):
                filename = "pacientes_valentina.xlsx"
                style_excel(df, filename)
                with open(filename, "rb") as file:
                    st.download_button("Descargar archivo Excel", data=file, file_name=filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.info("No hay pacientes registrados.")


    # ---------------- BIRTHDAYS ----------------
    elif page == "🎂 Cumpleaños":
        st.subheader("Cumpleaños del mes y del próximo mes")
        MONTHS_ES = {
            "January": "enero", "February": "febrero", "March": "marzo",
            "April": "abril", "May": "mayo", "June": "junio",
            "July": "julio", "August": "agosto", "September": "septiembre",
            "October": "octubre", "November": "noviembre", "December": "diciembre"
        }
        query = supabase.table("pacientes").select("*").neq("estado", "fallecido").execute()
        df = pd.DataFrame(query.data)
        if not df.empty:
            df["fecha_nacimiento"] = pd.to_datetime(df["fecha_nacimiento"], errors="coerce").dt.date
            current_month = datetime.today().month
            next_month = (current_month % 12) + 1
            df["Edad"] = df["fecha_nacimiento"].apply(calculate_age)
            df_this_month = df[pd.to_datetime(df["fecha_nacimiento"]).dt.month == current_month]
            df_next_month = df[pd.to_datetime(df["fecha_nacimiento"]).dt.month == next_month]


            current_month_name = MONTHS_ES[datetime.today().strftime('%B')]
            st.markdown(f"### 🎉 Cumpleaños de **{current_month_name}**")
            display_wrapped_table(df_this_month[["nombre", "fecha_nacimiento", "Edad", "estado"]])


            next_month_name_en = datetime(datetime.today().year, next_month, 1).strftime('%B')
            next_month_name = MONTHS_ES[next_month_name_en]
            st.markdown(f"### 🎈 Cumpleaños de **{next_month_name}**")
            display_wrapped_table(df_next_month[["nombre", "fecha_nacimiento", "Edad", "estado"]])
        else:
            st.info("No hay pacientes registrados.")


    # ---------------- FOOTER ----------------
    last_user, last_time = get_last_edit()
    if last_user and last_time:
        try:
            formatted_time = datetime.fromisoformat(last_time).strftime('%d/%m/%Y %H:%M')
        except Exception:
            formatted_time = last_time
        st.sidebar.markdown(
            f"<div class='bottom-left'>Última edición por <b>{last_user}</b> el {formatted_time}</div>",
            unsafe_allow_html=True
        )

